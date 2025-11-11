"""
Employee Excel Import/Export Router
Simple, dedicated router for Excel operations - NO AUTHENTICATION REQUIRED for download
"""

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from fastapi.responses import StreamingResponse
from typing import List
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import uuid
import random

from services.supabase_client import get_supabase_client
from utils.auth import hash_password, require_manager_or_admin
from models.user import User

# Create dedicated router for Excel operations
router = APIRouter()

@router.get("/download-template")
async def download_excel_template():
    """
    Download Excel template for employee import
    PUBLIC ENDPOINT - NO AUTHENTICATION REQUIRED
    
    Returns:
        Excel file with 6 sheets:
        1. Mẫu nhân viên (Template with dropdowns)
        2. Tra cứu nhanh (Quick reference - All mappings)
        3. Danh sách vai trò (Roles detail)
        4. Danh sách phòng ban (Departments detail)
        5. Danh sách chức vụ (Positions detail)
        6. Hướng dẫn (Instructions)
    """
    try:
        supabase = get_supabase_client()
        
        # Get departments and positions
        departments_result = supabase.table("departments").select("id, name, code").execute()
        departments = departments_result.data or []
        
        positions_result = supabase.table("positions").select("id, name, code, department_id").execute()
        positions = positions_result.data or []
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # ==================== SHEET 1: Template ====================
        ws_template = wb.create_sheet("Mẫu nhân viên", 0)
        
        headers = [
            "Họ *", "Tên *", "Email *", "Số điện thoại",
            "Mã phòng ban", "Mã chức vụ", "Ngày vào làm *",
            "Lương", "Vai trò *", "Mật khẩu"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws_template.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Example data
        if departments and positions:
            example = [
                "Nguyễn", "Văn A", "example1@company.com", "0901234567",
                departments[0]["code"], positions[0]["code"], 
                "2024-01-15", "15000000", "employee", "123456"
            ]
            for col, value in enumerate(example, start=1):
                ws_template.cell(row=2, column=col, value=value)
        
        # Column widths
        widths = [12, 12, 30, 15, 18, 18, 18, 15, 15, 12]
        for col, width in enumerate(widths, start=1):
            ws_template.column_dimensions[chr(64 + col)].width = width
        
        # ==================== SHEET 2: Quick Reference ====================
        ws_ref = wb.create_sheet("Tra cứu nhanh", 1)
        
        # Title
        ws_ref.merge_cells('A1:D1')
        title_cell = ws_ref.cell(1, 1, "📋 TRA CỨU NHANH - BẢNG ĐỐI CHIẾU MÃ VÀ TÊN")
        title_cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        title_cell.font = Font(bold=True, color="FFFFFF", size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_ref.row_dimensions[1].height = 30
        
        # Section 1: Departments
        current_row = 3
        ws_ref.cell(current_row, 1, "🏢 PHÒNG BAN").fill = header_fill
        ws_ref.cell(current_row, 1).font = header_font
        ws_ref.cell(current_row, 2, "").fill = header_fill
        ws_ref.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        ws_ref.cell(current_row, 1, "Mã phòng ban").fill = PatternFill(start_color="E7E9EB", end_color="E7E9EB", fill_type="solid")
        ws_ref.cell(current_row, 1).font = Font(bold=True)
        ws_ref.cell(current_row, 2, "Tên phòng ban").fill = PatternFill(start_color="E7E9EB", end_color="E7E9EB", fill_type="solid")
        ws_ref.cell(current_row, 2).font = Font(bold=True)
        current_row += 1
        
        for dept in departments:
            ws_ref.cell(current_row, 1, dept["code"])
            ws_ref.cell(current_row, 2, dept["name"])
            current_row += 1
        
        # Section 2: Positions
        current_row += 1
        ws_ref.cell(current_row, 1, "👔 CHỨC VỤ").fill = header_fill
        ws_ref.cell(current_row, 1).font = header_font
        ws_ref.cell(current_row, 2, "").fill = header_fill
        ws_ref.cell(current_row, 3, "").fill = header_fill
        ws_ref.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        ws_ref.cell(current_row, 1, "Mã chức vụ").fill = PatternFill(start_color="E7E9EB", end_color="E7E9EB", fill_type="solid")
        ws_ref.cell(current_row, 1).font = Font(bold=True)
        ws_ref.cell(current_row, 2, "Tên chức vụ").fill = PatternFill(start_color="E7E9EB", end_color="E7E9EB", fill_type="solid")
        ws_ref.cell(current_row, 2).font = Font(bold=True)
        ws_ref.cell(current_row, 3, "Thuộc phòng ban").fill = PatternFill(start_color="E7E9EB", end_color="E7E9EB", fill_type="solid")
        ws_ref.cell(current_row, 3).font = Font(bold=True)
        current_row += 1
        
        for pos in positions:
            ws_ref.cell(current_row, 1, pos["code"])
            ws_ref.cell(current_row, 2, pos["name"])
            dept_name = ""
            if pos.get("department_id"):
                for d in departments:
                    if d["id"] == pos["department_id"]:
                        dept_name = d["name"]
                        break
            ws_ref.cell(current_row, 3, dept_name)
            current_row += 1
        
        # Section 3: Roles
        current_row += 1
        ws_ref.cell(current_row, 1, "🎭 VAI TRÒ").fill = header_fill
        ws_ref.cell(current_row, 1).font = header_font
        ws_ref.cell(current_row, 2, "").fill = header_fill
        ws_ref.cell(current_row, 3, "").fill = header_fill
        ws_ref.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        ws_ref.cell(current_row, 1, "Mã vai trò").fill = PatternFill(start_color="E7E9EB", end_color="E7E9EB", fill_type="solid")
        ws_ref.cell(current_row, 1).font = Font(bold=True)
        ws_ref.cell(current_row, 2, "Tên vai trò").fill = PatternFill(start_color="E7E9EB", end_color="E7E9EB", fill_type="solid")
        ws_ref.cell(current_row, 2).font = Font(bold=True)
        ws_ref.cell(current_row, 3, "Mô tả").fill = PatternFill(start_color="E7E9EB", end_color="E7E9EB", fill_type="solid")
        ws_ref.cell(current_row, 3).font = Font(bold=True)
        current_row += 1
        
        roles = [
            ("admin", "Quản trị viên", "Quyền quản trị toàn hệ thống"),
            ("accountant", "Kế toán", "Quản lý tài chính và báo cáo"),
            ("sales", "Bán hàng", "Quản lý bán hàng và khách hàng"),
            ("workshop_employee", "Nhân viên xưởng", "Làm việc tại xưởng sản xuất"),
            ("employee", "Nhân viên", "Nhân viên văn phòng thông thường"),
            ("worker", "Công nhân", "Công nhân sản xuất"),
            ("transport", "Vận chuyển", "Nhân viên vận chuyển"),
            ("customer", "Khách hàng", "Tài khoản khách hàng")
        ]
        
        for code, name, desc in roles:
            ws_ref.cell(current_row, 1, code)
            ws_ref.cell(current_row, 2, name)
            ws_ref.cell(current_row, 3, desc)
            current_row += 1
        
        ws_ref.column_dimensions['A'].width = 22
        ws_ref.column_dimensions['B'].width = 30
        ws_ref.column_dimensions['C'].width = 35
        
        # ==================== SHEET 3: Roles (Detail) ====================
        ws_roles = wb.create_sheet("Danh sách vai trò")
        
        ws_roles.cell(1, 1, "Mã vai trò").fill = header_fill
        ws_roles.cell(1, 1).font = header_font
        ws_roles.cell(1, 2, "Tên vai trò").fill = header_fill
        ws_roles.cell(1, 2).font = header_font
        ws_roles.cell(1, 3, "Mô tả").fill = header_fill
        ws_roles.cell(1, 3).font = header_font
        
        for idx, (code, name, desc) in enumerate(roles, start=2):
            ws_roles.cell(idx, 1, code)
            ws_roles.cell(idx, 2, name)
            ws_roles.cell(idx, 3, desc)
        
        ws_roles.column_dimensions['A'].width = 25
        ws_roles.column_dimensions['B'].width = 25
        ws_roles.column_dimensions['C'].width = 40
        
        # ==================== SHEET 4: Departments ====================
        ws_dept = wb.create_sheet("Danh sách phòng ban")
        
        ws_dept.cell(1, 1, "Mã phòng ban").fill = header_fill
        ws_dept.cell(1, 1).font = header_font
        ws_dept.cell(1, 2, "Tên phòng ban").fill = header_fill
        ws_dept.cell(1, 2).font = header_font
        
        for idx, dept in enumerate(departments, start=2):
            ws_dept.cell(idx, 1, dept["code"])
            ws_dept.cell(idx, 2, dept["name"])
        
        ws_dept.column_dimensions['A'].width = 18
        ws_dept.column_dimensions['B'].width = 30
        
        # ==================== SHEET 5: Positions ====================
        ws_pos = wb.create_sheet("Danh sách chức vụ")
        
        ws_pos.cell(1, 1, "Mã chức vụ").fill = header_fill
        ws_pos.cell(1, 1).font = header_font
        ws_pos.cell(1, 2, "Tên chức vụ").fill = header_fill
        ws_pos.cell(1, 2).font = header_font
        ws_pos.cell(1, 3, "Thuộc phòng ban").fill = header_fill
        ws_pos.cell(1, 3).font = header_font
        
        for idx, pos in enumerate(positions, start=2):
            ws_pos.cell(idx, 1, pos["code"])
            ws_pos.cell(idx, 2, pos["name"])
            # Find department name
            dept_name = ""
            if pos.get("department_id"):
                for d in departments:
                    if d["id"] == pos["department_id"]:
                        dept_name = d["name"]
                        break
            ws_pos.cell(idx, 3, dept_name)
        
        ws_pos.column_dimensions['A'].width = 18
        ws_pos.column_dimensions['B'].width = 30
        ws_pos.column_dimensions['C'].width = 25
        
        # ==================== DATA VALIDATION (Dropdowns) ====================
        # Departments dropdown
        if departments:
            dept_codes = ",".join([d["code"] for d in departments])
            dv_dept = DataValidation(type="list", formula1=f'"{dept_codes}"', allow_blank=True)
            dv_dept.error = 'Chọn từ danh sách'
            dv_dept.errorTitle = 'Mã không hợp lệ'
            dv_dept.prompt = 'Click để chọn mã phòng ban'
            dv_dept.promptTitle = 'Chọn phòng ban'
            ws_template.add_data_validation(dv_dept)
            dv_dept.add('E2:E1000')
        
        # Positions dropdown
        if positions:
            pos_codes = ",".join([p["code"] for p in positions])
            dv_pos = DataValidation(type="list", formula1=f'"{pos_codes}"', allow_blank=True)
            dv_pos.error = 'Chọn từ danh sách'
            dv_pos.errorTitle = 'Mã không hợp lệ'
            dv_pos.prompt = 'Click để chọn mã chức vụ'
            dv_pos.promptTitle = 'Chọn chức vụ'
            ws_template.add_data_validation(dv_pos)
            dv_pos.add('F2:F1000')
        
        # Roles dropdown
        role_codes = "admin,accountant,sales,workshop_employee,employee,worker,transport,customer"
        dv_role = DataValidation(type="list", formula1=f'"{role_codes}"', allow_blank=False)
        dv_role.error = 'Chọn từ danh sách'
        dv_role.errorTitle = 'Vai trò không hợp lệ'
        dv_role.prompt = 'Click để chọn vai trò'
        dv_role.promptTitle = 'Chọn vai trò'
        ws_template.add_data_validation(dv_role)
        dv_role.add('I2:I1000')
        
        # ==================== SHEET 6: Instructions ====================
        ws_inst = wb.create_sheet("Hướng dẫn")
        
        instructions = [
            ["🎯 HƯỚNG DẪN NHẬP DỮ LIỆU", ""],
            ["", ""],
            ["📋 TRA CỨU NHANH", ""],
            ["Xem sheet 'Tra cứu nhanh'", "Bảng đối chiếu đầy đủ mã và tên"],
            ["", "Phòng ban + Chức vụ + Vai trò"],
            ["", ""],
            ["✅ CÁC CỘT BẮT BUỘC (*)", ""],
            ["Họ *", "Họ của nhân viên"],
            ["Tên *", "Tên của nhân viên"],
            ["Email *", "Email duy nhất, dùng để đăng nhập"],
            ["Ngày vào làm *", "Định dạng YYYY-MM-DD, ví dụ: 2024-01-15"],
            ["Vai trò *", "Chọn từ dropdown hoặc tra sheet 'Tra cứu nhanh'"],
            ["", ""],
            ["📝 CÁC CỘT TÙY CHỌN", ""],
            ["Số điện thoại", "Số điện thoại liên hệ"],
            ["Mã phòng ban", "Chọn từ dropdown hoặc tra sheet 'Tra cứu nhanh'"],
            ["Mã chức vụ", "Chọn từ dropdown hoặc tra sheet 'Tra cứu nhanh'"],
            ["Lương", "Số tiền (chỉ nhập số)"],
            ["Mật khẩu", "Mặc định 123456 (tối thiểu 6 ký tự)"],
            ["", ""],
            ["⚠️ LƯU Ý", ""],
            ["1. Email phải duy nhất", ""],
            ["2. Sử dụng dropdown", "Click vào ô để chọn"],
            ["3. Xem sheet 'Tra cứu nhanh'", "Để xem tên đầy đủ của mã"],
            ["4. Xóa dòng ví dụ", "Trước khi import"],
        ]
        
        for row_idx, (col1, col2) in enumerate(instructions, start=1):
            ws_inst.cell(row_idx, 1, col1)
            ws_inst.cell(row_idx, 2, col2)
        
        ws_inst.column_dimensions['A'].width = 35
        ws_inst.column_dimensions['B'].width = 50
        
        # ==================== SAVE AND RETURN ====================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=mau_nhap_nhan_vien.xlsx",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Không thể tạo file mẫu: {str(e)}"
        )


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(require_manager_or_admin)
):
    """
    Upload and import employees from Excel
    REQUIRES AUTHENTICATION: Admin or Manager only
    """
    try:
        # Validate file
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(400, "File phải là Excel (.xlsx hoặc .xls)")
        
        # Read file
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), sheet_name="Mẫu nhân viên")
        
        # Validate columns
        required = ["Họ *", "Tên *", "Email *", "Ngày vào làm *", "Vai trò *"]
        missing = [col for col in required if col not in df.columns]
        if missing:
            raise HTTPException(400, f"Thiếu cột: {', '.join(missing)}")
        
        supabase = get_supabase_client()
        
        # Get mappings
        depts = supabase.table("departments").select("id, code").execute()
        poss = supabase.table("positions").select("id, code").execute()
        
        dept_map = {d["code"]: d["id"] for d in depts.data or []}
        pos_map = {p["code"]: p["id"] for p in poss.data or []}
        
        # Get current user info for audit trail
        current_user_id = str(current_user.id)
        current_user_email = current_user.email
        
        # Process rows
        success = 0
        errors = []
        for idx, row in df.iterrows():
            row_num = idx + 2
            try:
                # Extract data
                first_name = str(row.get("Họ *", "")).strip()
                last_name = str(row.get("Tên *", "")).strip()
                email = str(row.get("Email *", "")).strip().lower()
                phone = str(row.get("Số điện thoại", "")).strip() if pd.notna(row.get("Số điện thoại")) else None
                dept_code = str(row.get("Mã phòng ban", "")).strip() if pd.notna(row.get("Mã phòng ban")) else None
                pos_code = str(row.get("Mã chức vụ", "")).strip() if pd.notna(row.get("Mã chức vụ")) else None
                hire_date_str = str(row.get("Ngày vào làm *", "")).strip()
                salary = row.get("Lương") if pd.notna(row.get("Lương")) else None
                role = str(row.get("Vai trò *", "employee")).strip().lower()
                password = str(row.get("Mật khẩu", "123456")).strip()
                
                # Validate
                if not all([first_name, last_name, email, hire_date_str]):
                    errors.append(f"Dòng {row_num}: Thiếu thông tin bắt buộc")
                    continue
                
                if "@" not in email:
                    errors.append(f"Dòng {row_num}: Email không hợp lệ")
                    continue
                
                # Parse date
                try:
                    hire_date = pd.to_datetime(hire_date_str).date().isoformat()
                except:
                    errors.append(f"Dòng {row_num}: Ngày không hợp lệ (dùng YYYY-MM-DD)")
                    continue
                
                # Check email exists
                existing = supabase.table("users").select("id").eq("email", email).execute()
                if existing.data:
                    errors.append(f"Dòng {row_num}: Email {email} đã tồn tại")
                    continue
                
                # Generate employee code
                emp_code = f"EMP{datetime.now().strftime('%Y%m')}{random.randint(1000, 9999)}"
                while supabase.table("employees").select("id").eq("employee_code", emp_code).execute().data:
                    emp_code = f"EMP{datetime.now().strftime('%Y%m')}{random.randint(1000, 9999)}"
                
                # Create auth user
                auth_resp = supabase.auth.admin.create_user({
                    "email": email,
                    "password": password,
                    "email_confirm": True,
                    "user_metadata": {
                        "full_name": f"{first_name} {last_name}",
                        "role": role
                    }
                })
                
                if not auth_resp.user:
                    errors.append(f"Dòng {row_num}: Không tạo được tài khoản")
                    continue
                
                user_id = auth_resp.user.id
                
                # Create user record
                supabase.table("users").insert({
                    "id": user_id,
                    "email": email,
                    "full_name": f"{first_name} {last_name}",
                    "role": role,
                    "password_hash": hash_password(password),
                    "is_active": True,
                    "created_by": current_user_id,
                    "updated_by": current_user_id,
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                }).execute()
                
                # Create employee
                supabase.table("employees").insert({
                    "id": str(uuid.uuid4()),
                    "user_id": user_id,
                    "employee_code": emp_code,
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email,
                    "phone": phone,
                    "department_id": dept_map.get(dept_code),
                    "position_id": pos_map.get(pos_code),
                    "hire_date": hire_date,
                    "salary": float(salary) if salary else None,
                    "status": "active",
                    "created_by": current_user_id,
                    "updated_by": current_user_id,
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                }).execute()
                
                success += 1
                
            except Exception as e:
                errors.append(f"Dòng {row_num}: {str(e)}")
                continue
        
        return {
            "message": "Hoàn thành import",
            "success_count": success,
            "error_count": len(errors),
            "total_rows": len(df),
            "imported_by": current_user_email,
            "imported_by_id": current_user_id,
            "errors": errors[:20]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Lỗi xử lý file: {str(e)}")

